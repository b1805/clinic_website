from flask import Flask, request, send_file, render_template
import pandas as pd
import io
from openpyxl import load_workbook
import discord
from discord.ext import commands
from dotenv import load_dotenv
import os
import asyncio
import zipfile
import tempfile
import re
import io
from difflib import get_close_matches
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from io import BytesIO
from collections import Counter

app = Flask(__name__)

try:
    avonet_path = 'AVONET Supplementary dataset 1.xlsx'
    avonet_sheets = pd.read_excel(avonet_path, sheet_name=None)
    avonet_df = pd.concat(avonet_sheets.values(), ignore_index=True)
    
    anage_df = pd.read_csv('anage_data.txt', delimiter='\t', low_memory=False)
    
    ParentalProvisioning_path = 'ParentalProvisioning.xlsx'
    ParentalProvisioning_sheets = pd.read_excel(ParentalProvisioning_path, sheet_name=None)
    ParentalProvisioning_df = pd.concat(ParentalProvisioning_sheets.values(), ignore_index=True)
    
    BrainSize1_path = 'Data Sheet 1.xlsx'
    BrainSize2_path = 'Data Sheet 2.xlsx'
    BrainSize3_path = 'Data Sheet 3.xlsx'
    BrainSize4_path = 'Data Sheet 4.xlsx'
    BrainSize1_sheets = pd.read_excel(BrainSize1_path, sheet_name=None)
    BrainSize2_sheets = pd.read_excel(BrainSize2_path, sheet_name=None)
    BrainSize3_sheets = pd.read_excel(BrainSize3_path, sheet_name=None)
    BrainSize4_sheets = pd.read_excel(BrainSize4_path, sheet_name=None)
    BrainSize1_df = pd.concat(BrainSize1_sheets.values(), ignore_index=True)
    BrainSize2_df = pd.concat(BrainSize2_sheets.values(), ignore_index=True)
    BrainSize3_df = pd.concat(BrainSize3_sheets.values(), ignore_index=True)
    BrainSize4_df = pd.concat(BrainSize4_sheets.values(), ignore_index=True)
    
    NestTraitGlobal_path = 'Global-database-NestTrait_v2.csv'
    NestTraitGlobal_sheets = pd.read_csv(NestTraitGlobal_path)
    NestTraitGlobal_df = pd.concat([NestTraitGlobal_sheets], ignore_index=True)

    NestTraitsS1_path = 'Nest-traits-Dataset-S1.csv'
    NestTraitsS1_sheets = pd.read_csv(NestTraitsS1_path)
    NestTraitsS1_df = pd.concat([NestTraitsS1_sheets], ignore_index=True)

    NestTraitsS2_path = 'Nest-traits-Dataset-S2.csv'
    NestTraitsS2_sheets = pd.read_csv(NestTraitsS2_path, encoding='latin1')
    NestTraitsS2_df = pd.concat([NestTraitsS2_sheets], ignore_index=True)

    LatitudinalGradient_path = 'The latitudinal gradient-rawdata.csv'
    LatitudinalGradient_sheets = pd.read_csv(LatitudinalGradient_path, encoding='latin1')
    LatitudinalGradient_df = pd.concat([LatitudinalGradient_sheets], ignore_index=True)
    
    alt_strat_path = 'Alternative ecological strategies.csv'
    alt_strat_df = pd.read_csv(alt_strat_path, encoding='latin1')

    aoh_path = 'Area of Habitat.csv'
    aoh_df = pd.read_csv(aoh_path)

    beak_shape_path = 'Beak shape.xlsx'
    beak_shape_sheets = pd.read_excel(beak_shape_path, sheet_name=None)
    beak_shape_df = pd.concat(beak_shape_sheets.values(), ignore_index=True)

    clutch1_path = 'clutch_dataset1.csv'
    clutch1_df = pd.read_csv(clutch1_path)

    clutch2_path = 'clutch_dataset2.csv'
    clutch2_df = pd.read_csv(clutch2_path)

    savitrait1_path = 'SAviTraits_1-0_1.csv'
    savitrait1_df = pd.read_csv(savitrait1_path, encoding='latin1')

    savitrait2_path = 'SAviTraits_1-0_2.csv'
    savitrait2_df = pd.read_csv(savitrait2_path)

    # a_path = 'a.xlsx'
    # a_sheets = pd.read_excel(a_path, sheet_name=None)
    # a_df = pd.concat(a_sheets.values(), ignore_index=True)
    print("✅ Datasets loaded successfully.")
except FileNotFoundError:
    print("❌ Dataset files not found. Please ensure they are in the correct location.")
    avonet_df = None
    anage_df = None
    ParentalProvisioning_df = None
    BrainSize1_df = None
    BrainSize2_df = None
    BrainSize3_df = None
    BrainSize4_df = None
    NestTraitGlobal_df = None
    NestTraitsS1_df = None
    NestTraitsS2_df = None
    LatitudinalGradient_df = None
    #a_df = None

# Your column merge map
column_merge_map = {
                "Source": ["Source"],
                "species": ["Species1", "Species2", "Species3", "pigot_Species3", "Species", "Species scientific name", "Scientific_name", "Unique_Scientific_Name", "tip_label", "BrainSize1_species", "BrainSize2_species", "BrainSize3_species", "BrainSize4_species", "Species scientific name 2", "Species_Scientific_Name", "Species_Scientific_Name_2", "Alt_Species", "BINOMIAL", "PhyloName", "Binomial", "ebird_name"],
                "common_name": ["Common name", "Species common name", "English_Name"],
                "family": ["Family1", "Family2", "Family3", "Family", "Family.name", "latinFamily"],
                "order": ["Order1", "Order2", "Order3", "Order", "Order_"],
                "genus": ["Genus"],
                "beak_length_mm": ["Beak.Length_Culmen", "Bill_TotalCulmen", "Beak.Length_Nares", "Bill_Nares"],
                "beak_width_mm": ["Beak.Width", "Bill_Width"],
                "beak_depth_mm": ["Beak.Depth", "Bill_Depth"],
                "tarsus_length_mm": ["Tarsus.Length", "Tarsus_Length"],
                "wing_length_mm": ["Wing.Length", "Wing_Chord"],
                "kipps_distance_mm": ["Kipps.Distance", "Kipp's_Distance"],
                "tail_length_mm": ["Tail.Length", "Tail_Length"],
                "mass_g": ["Mass", "Adult weight (g)", "Body mass (g)", "body.g", "Body.Mass.g"],
                "mass_log": ["Body mass (log)"],
                "brain_volume_mm3": ["Endocast volume (mm3)", "brain.ml"],
                "brain_mass_g": ["Brain.Mass.g"],
                "relative_brain_size": ["BrainResidual", "ReltiveBrainSive.OLSresiduals", "RelativeBrainSize.PGLSresiduals"],
                "hand_wing_index": ["Hand-Wing.Index", "Hand-wing.Index", "Hand-Wing Index (Claramunt 2011)", "HWI"],
                "range_size_km2": ["Range.Size"],
                "latitude": ["Latitude", "Centroid.Latitude"],
                "longitude": ["Longitude", "Centroid.Longitude"],
                "migration_status": ["Migration", "MigratoryBehaviour", "is.migratory", "migration"],
                "habitat_type": ["Habitat", "habitat", "AOH_name"],
                "habitat_density": ["Habitat.Density", "hab-dens"],
                "trophic_level": ["Trophic.Level", "trophiclevel"],
                "trophic_niche": ["Trophic.Niche", "niche"],
                "primary_lifestyle": ["Primary.Lifestyle", "lifestyle"],
                "egg_mass_g": ["egg_mass"],
                "clutch_size": ["Litter/Clutch size", "clutch_size", "Eggs/year"],
                "broods_per_year": ["Litters/Clutches per year", "Broods"],
                "birth_weight_g": ["Birth weight (g)"],
                "weaning_weight_g": ["Weaning weight (g)"],
                "growth_rate": ["Growth rate (1/days)"],
                "longevity_years": ["Maximum longevity (yrs)", "LifespanMax", "longevity"],
                "maturity_days_female": ["Female maturity (days)"],
                "maturity_days_male": ["Male maturity (days)"],
                "gestation_incubation_days": ["Gestation/Incubation (days)"],
                "weaning_days": ["Weaning (days)"],
                "metabolic_rate_W": ["Metabolic rate (W)"],
                "temperature_K": ["Temperature (K)"],
                "urban_tolerance_index": ["UrbanToleranceIndex"],
                "urban_abundance": ["AbundaceUrban", "AverageUrbanAbundance"],
                "wild_abundance": ["AbundanceWild", "AverageWildAbundance"],
                "social_bonds": ["social_bonds"],
                "caretakers": ["caretakers"],
                "diet_category": ["Diet_Cat"],
                "diet_subcategory": ["Diet_Sub_Cat"],
                "diet_variability": ["Diet_Variability"],
                "foraging_strategy": ["foraging"],
                "sedentariness": ["sedentariness"],
                "insularity": ["insularity"],
                "grouping_behavior": ["grouping"]
            }

@app.route('/')
def index():
    return render_template('bot.html', traits=column_merge_map.keys())

@app.route('/generate', methods=['POST'])
def generate():
    species_name = request.form['species']
    user_selected_new_cols = request.form.getlist('traits')

    def get_bird_data_sync(species_name):
            avonet_matches_idx = set()
            species_words = species_name.lower().split()

            def matches_all_words(cell_value: str) -> bool:
                # If the user requests all data, return True for every row
                if species_name.strip().upper() == "ALL DATA":
                    return True

                cell_value = str(cell_value).lower()
                cell_words = cell_value.replace('_', ' ').split()

                # Generate two forms of species input: with spaces and with underscores replaced
                species_words_normal = species_name.lower().split()
                species_words_underscore = species_name.lower().replace(' ', '_').split('_')

                def match_words(target_words):
                    if len(target_words) == 1:
                        return any(get_close_matches(target_words[0], [cw], n=1, cutoff=0.85) for cw in cell_words)
                    return all(any(get_close_matches(sw, [cw], n=1, cutoff=0.85) for cw in cell_words) for sw in target_words)

                return match_words(species_words_normal) or match_words(species_words_underscore)


            for col in avonet_df.columns:
                matches = avonet_df[col].dropna().astype(str).apply(matches_all_words)
                matched_rows = matches[matches].index
                avonet_matches_idx.update(matched_rows)

            avonet_data = avonet_df.loc[list(avonet_matches_idx)]
            
            # Search AnAge: same logic + enhanced search for combined genus + species
            anage_matches_idx = set()

            # Check if species_name is more than one word
            species_words = species_name.strip().split()
            if len(species_words) > 1 and 'Genus' in anage_df.columns and 'Species' in anage_df.columns:
                # Create a combined column of 'Genus' + 'Species'
                combined_genus_species = (anage_df['Genus'].astype(str) + ' ' + anage_df['Species'].astype(str)).dropna()

                # Search the combined column
                combined_matches = combined_genus_species.astype(str).apply(matches_all_words)
                combined_matched_rows = combined_matches[combined_matches].index
                anage_matches_idx.update(combined_matched_rows)

            # Continue with the original column-wise search
            for col in anage_df.columns:
                matches = anage_df[col].dropna().astype(str).apply(matches_all_words)
                matched_rows = matches[matches].index
                anage_matches_idx.update(matched_rows)

            anage_data = anage_df.loc[list(anage_matches_idx)]

            if avonet_data.empty and anage_data.empty:
                return None

            if not avonet_data.empty:
                avonet_data = avonet_data.copy()
                avonet_data['Source'] = 'AVONET'

            if not anage_data.empty:
                anage_data = anage_data.copy()
                anage_data['Source'] = 'AnAge'
                
            # === ParentalProvisioning Dataset Search ===
            ParentalProvisioning_matches_idx = set()
            for col in ParentalProvisioning_df.columns:
                matches = ParentalProvisioning_df[col].dropna().astype(str).apply(matches_all_words)
                ParentalProvisioning_matches_idx.update(matches[matches].index)
            ParentalProvisioning_data = ParentalProvisioning_df.loc[list(ParentalProvisioning_matches_idx)]
            if not ParentalProvisioning_data.empty:
                ParentalProvisioning_data = ParentalProvisioning_data.copy()
                ParentalProvisioning_data['Source'] = 'ParentalProvisioning'
                
            # === 'BrainSize1' Dataset Search ===
            BrainSize1_matches_idx = set()
            for col in BrainSize1_df.columns:
                matches = BrainSize1_df[col].dropna().astype(str).apply(matches_all_words)
                BrainSize1_matches_idx.update(matches[matches].index)
            BrainSize1_data = BrainSize1_df.loc[list(BrainSize1_matches_idx)]
            if not BrainSize1_data.empty:
                BrainSize1_data = BrainSize1_data.copy()
                BrainSize1_data['Source'] = 'BrainSize1'
                
            # === 'BrainSize2' Dataset Search ===
            BrainSize2_matches_idx = set()
            for col in BrainSize2_df.columns:
                matches = BrainSize2_df[col].dropna().astype(str).apply(matches_all_words)
                BrainSize2_matches_idx.update(matches[matches].index)
            BrainSize2_data = BrainSize2_df.loc[list(BrainSize2_matches_idx)]
            if not BrainSize2_data.empty:
                BrainSize2_data = BrainSize2_data.copy()
                BrainSize2_data['Source'] = 'BrainSize2'
                
            # === 'BrainSize3' Dataset Search ===
            BrainSize3_matches_idx = set()
            for col in BrainSize3_df.columns:
                matches = BrainSize3_df[col].dropna().astype(str).apply(matches_all_words)
                BrainSize3_matches_idx.update(matches[matches].index)
            BrainSize3_data = BrainSize3_df.loc[list(BrainSize3_matches_idx)]
            if not BrainSize3_data.empty:
                BrainSize3_data = BrainSize3_data.copy()
                BrainSize3_data['Source'] = 'BrainSize3'
                
            # === 'BrainSize4' Dataset Search ===
            BrainSize4_matches_idx = set()
            for col in BrainSize4_df.columns:
                matches = BrainSize4_df[col].dropna().astype(str).apply(matches_all_words)
                BrainSize4_matches_idx.update(matches[matches].index)
            BrainSize4_data = BrainSize4_df.loc[list(BrainSize4_matches_idx)]
            if not BrainSize4_data.empty:
                BrainSize4_data = BrainSize4_data.copy()
                BrainSize4_data['Source'] = 'BrainSize4'
                
            # === NestTraitGlobal Dataset Search ===
            NestTraitGlobal_matches_idx = set()
            for col in NestTraitGlobal_df.columns:
                matches = NestTraitGlobal_df[col].dropna().astype(str).apply(matches_all_words)
                NestTraitGlobal_matches_idx.update(matches[matches].index)
            NestTraitGlobal_data = NestTraitGlobal_df.loc[list(NestTraitGlobal_matches_idx)]
            if not NestTraitGlobal_data.empty:
                NestTraitGlobal_data = NestTraitGlobal_data.copy()
                NestTraitGlobal_data['Source'] = 'NestTraitGlobal'

            # === NestTraitsS1 Dataset Search ===
            NestTraitsS1_matches_idx = set()
            for col in NestTraitsS1_df.columns:
                matches = NestTraitsS1_df[col].dropna().astype(str).apply(matches_all_words)
                NestTraitsS1_matches_idx.update(matches[matches].index)
            NestTraitsS1_data = NestTraitsS1_df.loc[list(NestTraitsS1_matches_idx)]
            if not NestTraitsS1_data.empty:
                NestTraitsS1_data = NestTraitsS1_data.copy()
                NestTraitsS1_data['Source'] = 'NestTraitsS1'

            # === NestTraitsS2 Dataset Search ===
            NestTraitsS2_matches_idx = set()
            for col in NestTraitsS2_df.columns:
                matches = NestTraitsS2_df[col].dropna().astype(str).apply(matches_all_words)
                NestTraitsS2_matches_idx.update(matches[matches].index)
            NestTraitsS2_data = NestTraitsS2_df.loc[list(NestTraitsS2_matches_idx)]
            if not NestTraitsS2_data.empty:
                NestTraitsS2_data = NestTraitsS2_data.copy()
                NestTraitsS2_data['Source'] = 'NestTraitsS2'
                
            # === 'LatitudinalGradient' Dataset Search ===
            LatitudinalGradient_matches_idx = set()
            for col in LatitudinalGradient_df.columns:
                matches = LatitudinalGradient_df[col].dropna().astype(str).apply(matches_all_words)
                LatitudinalGradient_matches_idx.update(matches[matches].index)
            LatitudinalGradient_data = LatitudinalGradient_df.loc[list(LatitudinalGradient_matches_idx)]
            if not LatitudinalGradient_data.empty:
                LatitudinalGradient_data = LatitudinalGradient_data.copy()
                LatitudinalGradient_data['Source'] = 'LatitudinalGradient'
                
            # === Alternative Ecological Strategies Search ===
            alt_strat_matches_idx = set()
            for col in alt_strat_df.columns:
                matches = alt_strat_df[col].dropna().astype(str).apply(matches_all_words)
                alt_strat_matches_idx.update(matches[matches].index)
            alt_strat_data = alt_strat_df.loc[list(alt_strat_matches_idx)]
            if not alt_strat_data.empty:
                alt_strat_data = alt_strat_data.copy()
                alt_strat_data['Source'] = 'Alternative ecological strategies'

            # === Area of Habitat Search ===
            aoh_matches_idx = set()
            for col in aoh_df.columns:
                matches = aoh_df[col].dropna().astype(str).apply(matches_all_words)
                aoh_matches_idx.update(matches[matches].index)
            aoh_data = aoh_df.loc[list(aoh_matches_idx)]
            if not aoh_data.empty:
                aoh_data = aoh_data.copy()
                aoh_data['Source'] = 'Area of Habitat'

            # === Beak Shape Search ===
            beak_shape_matches_idx = set()
            for col in beak_shape_df.columns:
                matches = beak_shape_df[col].dropna().astype(str).apply(matches_all_words)
                beak_shape_matches_idx.update(matches[matches].index)
            beak_shape_data = beak_shape_df.loc[list(beak_shape_matches_idx)]
            if not beak_shape_data.empty:
                beak_shape_data = beak_shape_data.copy()
                beak_shape_data['Source'] = 'Beak shape'

            # === Clutch Dataset 1 Search ===
            clutch1_matches_idx = set()
            for col in clutch1_df.columns:
                matches = clutch1_df[col].dropna().astype(str).apply(matches_all_words)
                clutch1_matches_idx.update(matches[matches].index)
            clutch1_data = clutch1_df.loc[list(clutch1_matches_idx)]
            if not clutch1_data.empty:
                clutch1_data = clutch1_data.copy()
                clutch1_data['Source'] = 'clutch_dataset1'

            # === Clutch Dataset 2 Search ===
            clutch2_matches_idx = set()
            for col in clutch2_df.columns:
                matches = clutch2_df[col].dropna().astype(str).apply(matches_all_words)
                clutch2_matches_idx.update(matches[matches].index)
            clutch2_data = clutch2_df.loc[list(clutch2_matches_idx)]
            if not clutch2_data.empty:
                clutch2_data = clutch2_data.copy()
                clutch2_data['Source'] = 'clutch_dataset2'

            # === SAviTraits 1-0 1 Search ===
            savitrait1_matches_idx = set()
            for col in savitrait1_df.columns:
                matches = savitrait1_df[col].dropna().astype(str).apply(matches_all_words)
                savitrait1_matches_idx.update(matches[matches].index)
            savitrait1_data = savitrait1_df.loc[list(savitrait1_matches_idx)]
            if not savitrait1_data.empty:
                savitrait1_data = savitrait1_data.copy()
                savitrait1_data['Source'] = 'SAviTraits_1-0_1'

            # === SAviTraits 1-0 2 Search ===
            savitrait2_matches_idx = set()
            for col in savitrait2_df.columns:
                matches = savitrait2_df[col].dropna().astype(str).apply(matches_all_words)
                savitrait2_matches_idx.update(matches[matches].index)
            savitrait2_data = savitrait2_df.loc[list(savitrait2_matches_idx)]
            if not savitrait2_data.empty:
                savitrait2_data = savitrait2_data.copy()
                savitrait2_data['Source'] = 'SAviTraits_1-0_2'


                
            # # === 'a' Dataset Search ===
            # a_matches_idx = set()
            # for col in a_df.columns:
            #     matches = a_df[col].dropna().astype(str).apply(matches_all_words)
            #     a_matches_idx.update(matches[matches].index)
            # a_data = a_df.loc[list(a_matches_idx)]
            # if not a_data.empty:
            #     a_data = a_data.copy()
            #     a_data['Source'] = 'A'

            
            # Combine all found dataframes vertically first (ignore Source columns here)
            dfs_to_combine = [df for df in [
                avonet_data, anage_data, ParentalProvisioning_data,
                BrainSize1_data, BrainSize2_data, BrainSize3_data, BrainSize4_data,
                NestTraitGlobal_data, NestTraitsS1_data, NestTraitsS2_data,
                LatitudinalGradient_data, alt_strat_data, aoh_data,
                beak_shape_data, clutch1_data, clutch2_data,
                savitrait1_data, savitrait2_data
            ] if df is not None and not df.empty]

            if not dfs_to_combine:
                return None

            combined_df = pd.concat(dfs_to_combine, ignore_index=True, sort=False)

            # Create a new dataframe with reduced columns
            merged_df = pd.DataFrame(index=combined_df.index)

            for new_col, old_cols in column_merge_map.items():
                if new_col not in user_selected_new_cols:
                    continue  # Skip columns not selected by the user
                
                existing_cols = [col for col in old_cols if col in combined_df.columns]
                if existing_cols:
                    filled = combined_df[existing_cols].bfill(axis=1).iloc[:, 0]
                    filled = filled.infer_objects(copy=False)
                    merged_df[new_col] = filled
                else:
                    merged_df[new_col] = None


            output = io.BytesIO()
            
            # Drop rows where all columns after column F are empty
            merged_df = merged_df[~merged_df.iloc[:, 6:].isnull().all(axis=1)]
            
            # Remove exact duplicate rows
            merged_df = merged_df.drop_duplicates()
            
            # Combine genus and species if genus is not null/empty
            merged_df["species"] = merged_df.apply(
                lambda row: f"{row['genus']} {row['species']}" if pd.notnull(row["genus"]) and str(row["genus"]).strip() != "" else row["species"],
                axis=1
            )
            
            # Drop underscores from species names before sorting
            merged_df["species"] = merged_df["species"].str.replace("_", " ", regex=False)
            
            # Sort the DataFrame by the "species" column before writing to Excel
            merged_df = merged_df.sort_values(by="species", ascending=True)

            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                merged_df.to_excel(writer, index=False)
            output.seek(0)

            # Load workbook and apply header styles
            wb = load_workbook(output)
            ws = wb.active

            # Save to final output
            final_output = io.BytesIO()
            wb.save(final_output)
            final_output.seek(0)
            return send_file(final_output, download_name=f'{species_name}_data.xlsx', as_attachment=True)
    result_file = get_bird_data_sync(species_name)
    return result_file

if __name__ == '__main__':
    app.run(debug=True)
